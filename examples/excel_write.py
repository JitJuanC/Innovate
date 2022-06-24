import openpyxl

def create_workbook(target_date, star_dict, feedback_list):
    wb = openpyxl.Workbook() # Create new excel workbook
    filename = "today_feedback.xlsx" # Workbook name
    # filename = "test.xlsx" # Workbook name

    ws = wb.active # Activate the worksheet of the workbook

     # The Content for this Worksheet
    ws['a1'] = "Date"
    ws['b1'] = str(target_date)
    ws['a2'] = "Star Ratings"
    for i, (star, count) in enumerate(star_dict.items(), 3):
        ws[f'a{i}'] = f"{star} Stars"
        ws[f'b{i}'] = count
        if star == 1:
            ws[f'a{i}'] = f"{star} Star"

    ws['a9'] = "Performance Count(s) Received Today"
    ws['a10'] = "The chatbot experience (e.g. intuitive)"
    ws['a11'] = "Clarity of the announcement (e.g. volume)"
    ws['a12'] = "Ease of understanding the FAQ page (e.g. comprehensive)"
    ws['a13'] = "Navigation feature (e.g. directing me to the appropriate machine)"
    for i, sum in enumerate(feedback_list[:4],10):
        ws[f'b{i}'] = sum
    
    ws['a15'] = "Features Count(s) Received Today"
    ws['a16'] = "Video call with staff for General Enquiry"
    ws['a17'] = "Show locations of nearby branches & ATM"
    ws['a18'] = "Notify customers of their unattended items (e.g. Wallet, Bag)"
    ws['a19'] = "Step-by-step guide to perform a particular transaction (e.g. Top-up of Ezlink card)"
    for i, sum in enumerate(feedback_list[-4:],16):
        ws[f'b{i}'] = sum
    
    wb.save(filename)

if __name__=='__main__':
    import MySQLdb

    def get_stars(sending_date, i):
        cursor.execute(f"select count(stars) from feedback where time like '{sending_date}%' and stars like {i}")
        star = cursor.fetchall()
        return star[0][0]

    db = MySQLdb.connect('172.18.0.3', 'username', 'password', db='dbs')
    cursor = db.cursor()
    target_date = '2021-10-07'
    cursor.execute(f"select cast(avg(f.stars) as decimal(10,1)), sum(p.intuitive), sum(p.volume), sum(p.comprehensive), \
                    sum(p.directing), sum(ft.video_call), sum(ft.show_location), sum(ft.unattended), sum(ft.transaction) \
                    from feedback f join performance p on f.id=p.id join features ft on f.id=ft.id where f.time like '{target_date}%'")
    fetched = cursor.fetchall() 
    stars = {i : get_stars(target_date, i) for i in range(1, 6)} 
    feedback_sum = [int(i) for i in fetched[0][1:]]
    create_workbook(target_date, stars, feedback_sum)
